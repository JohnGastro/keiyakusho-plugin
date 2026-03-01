#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fill_template.py - Rabinet Excel Template Filler

ラビーネットのExcelテンプレートに対して、契約情報を流し込むスクリプト。
openpyxlを使用してセル単位でデータを入力し、日付形式の変換や計算式の自動化を行う。

使用方法:
    python fill_template.py <template_path> <data_json> <output_path>

例:
    python fill_template.py "/path/to/template.xlsx" "{...config...}" "/path/to/output.xlsx"
"""

import json
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("エラー: openpyxlがインストールされていません")
    print("pip install openpyxl を実行してください")
    sys.exit(1)


class RabbynetContractFiller:
    """ラビーネット契約書テンプレート記入自動化クラス"""

    def __init__(self, template_path):
        """
        初期化

        Args:
            template_path (str): Excelテンプレートファイルのパス
        """
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f"テンプレートファイルが見つかりません: {template_path}")

        self.wb = load_workbook(self.template_path)
        self.ws = self.wb.active
        self.errors = []

    def convert_to_japanese_date(self, date_obj, format_style='reiwa'):
        """
        日付をPython datetime または 'YYYY-MM-DD' 文字列から日本語形式に変換

        Args:
            date_obj: datetime オブジェクトまたは 'YYYY-MM-DD' 形式の文字列
            format_style (str): 'reiwa' (令和X年X月X日) または 'slash' (R06/3/1)

        Returns:
            str: 日本語形式の日付文字列
        """
        try:
            # 文字列の場合はパース
            if isinstance(date_obj, str):
                date_obj = datetime.strptime(date_obj, '%Y-%m-%d')
            elif not isinstance(date_obj, datetime):
                raise ValueError("日付はdatetimeオブジェクトまたは'YYYY-MM-DD'形式の文字列である必要があります")

            year = date_obj.year
            month = date_obj.month
            day = date_obj.day

            if format_style == 'reiwa':
                # 令和形式: 令和X年X月X日
                if year >= 2019:
                    reiwa_year = year - 2018  # 2019 = 令和1年
                    return f"令和{reiwa_year}年{month}月{day}日"
                elif year >= 1989:
                    heisei_year = year - 1988  # 1989 = 平成1年
                    return f"平成{heisei_year}年{month}月{day}日"
                else:
                    return f"{year}年{month}月{day}日"
            elif format_style == 'slash':
                # スラッシュ形式: R06/3/1 (令和06/3/1)
                if year >= 2019:
                    reiwa_year = year - 2018
                    return f"R{reiwa_year:02d}/{month}/{day}"
                else:
                    return f"{year}/{month}/{day}"
            else:
                raise ValueError(f"未対応の日付形式: {format_style}")
        except Exception as e:
            raise ValueError(f"日付変換エラー: {str(e)}")

    def format_currency(self, value):
        """
        数値を通貨形式（カンマ区切り）に変換

        Args:
            value: 数値（int, float, または数値文字列）

        Returns:
            str: 通貨形式の文字列（例: '30,000,000'）
        """
        try:
            if isinstance(value, str):
                value = float(value.replace(',', ''))
            return f"{int(value):,}"
        except (ValueError, TypeError) as e:
            raise ValueError(f"通貨フォーマットエラー: {str(e)}")

    def fill_cell(self, cell_ref, value, cell_type='text', **kwargs):
        """
        セルに値を入力

        Args:
            cell_ref (str): セル位置（例: 'B5', 'C10'）
            value: 入力値
            cell_type (str): セルタイプ ('text', 'number', 'currency', 'date', 'formula', 'checkbox')
            **kwargs: 追加オプション（例: date_format='reiwa'）

        Returns:
            bool: 成功時True、失敗時False
        """
        try:
            cell = self.ws[cell_ref]

            if cell_type == 'text':
                cell.value = str(value)

            elif cell_type == 'number':
                cell.value = float(value)

            elif cell_type == 'currency':
                # 数値として入力し、セルに通貨フォーマット適用
                cell.value = float(value)
                cell.number_format = '#,##0'

            elif cell_type == 'date':
                date_format = kwargs.get('date_format', 'reiwa')
                cell.value = self.convert_to_japanese_date(value, date_format)

            elif cell_type == 'formula':
                # 計算式を入力（例: '=SUM(A1:A10)'）
                cell.value = value

            elif cell_type == 'checkbox':
                # チェックボックス: True/False → ○/空白
                if value:
                    cell.value = '○'
                else:
                    cell.value = ''

            else:
                raise ValueError(f"未対応のセルタイプ: {cell_type}")

            return True

        except Exception as e:
            error_msg = f"セル{cell_ref}への入力に失敗: {str(e)}"
            self.errors.append(error_msg)
            return False

    def fill_from_config(self, config):
        """
        設定辞書に基づいてセルを一括入力

        Args:
            config (dict): セルマッピング設定
                {
                    "contract_date": {
                        "cell": "C3",
                        "value": "2025-03-01",
                        "type": "date",
                        "date_format": "reiwa"
                    },
                    "seller_name": {
                        "cell": "B5",
                        "value": "田中太郎",
                        "type": "text"
                    },
                    ...
                }

        Returns:
            dict: 結果サマリー
        """
        success_count = 0
        failure_count = 0

        for field_name, field_config in config.items():
            if 'cell' not in field_config or 'value' not in field_config:
                self.errors.append(f"設定エラー {field_name}: 'cell' と 'value' が必須です")
                failure_count += 1
                continue

            cell_ref = field_config['cell']
            value = field_config['value']
            cell_type = field_config.get('type', 'text')

            # value が None または空文字列の場合はスキップ
            if value is None or value == '':
                continue

            # 追加のキーワード引数を抽出
            extra_kwargs = {k: v for k, v in field_config.items()
                          if k not in ['cell', 'value', 'type']}

            if self.fill_cell(cell_ref, value, cell_type, **extra_kwargs):
                success_count += 1
            else:
                failure_count += 1

        return {
            'success': success_count,
            'failure': failure_count,
            'total': success_count + failure_count,
            'errors': self.errors
        }

    def calculate_stamp_duty(self, contract_amount):
        """
        印紙税を計算（売買代金ベース）

        Args:
            contract_amount (float): 売買代金

        Returns:
            int: 印紙税額（円）
        """
        if contract_amount < 10_000:
            return 0
        elif contract_amount < 100_000:
            return 200
        elif contract_amount < 1_000_000:
            return 200
        elif contract_amount < 10_000_000:
            return 1_000
        elif contract_amount < 50_000_000:
            return 10_000
        elif contract_amount < 100_000_000:
            return 20_000
        elif contract_amount < 500_000_000:
            return 60_000
        else:
            return 100_000

    def calculate_brokerage_fee(self, contract_amount):
        """
        仲介手数料を計算（標準額）

        Args:
            contract_amount (float): 売買代金（または賃料年額）

        Returns:
            int: 仲介手数料（税抜）
        """
        # 標準額: 200万円以下は5%, 200万円超400万円以下は4%, 400万円超は3%
        if contract_amount <= 2_000_000:
            return int(contract_amount * 0.05)
        elif contract_amount <= 4_000_000:
            return int(2_000_000 * 0.05 + (contract_amount - 2_000_000) * 0.04)
        else:
            return int(2_000_000 * 0.05 + 2_000_000 * 0.04 + (contract_amount - 4_000_000) * 0.03)

    def save(self, output_path):
        """
        編集後のブックを保存

        Args:
            output_path (str): 出力先ファイルパス

        Returns:
            bool: 成功時True
        """
        try:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            self.wb.save(str(output_path))
            return True
        except Exception as e:
            self.errors.append(f"ファイル保存エラー: {str(e)}")
            return False

    def get_summary(self):
        """
        処理結果のサマリーを取得

        Returns:
            dict: サマリー情報
        """
        return {
            'template': str(self.template_path),
            'errors': self.errors,
            'error_count': len(self.errors)
        }


def load_config_from_json(json_str):
    """
    JSON文字列から設定を読み込む

    Args:
        json_str (str): JSON形式の設定文字列

    Returns:
        dict: パースされた設定
    """
    try:
        return json.loads(json_str)
    except json.JSONDecodeError as e:
        print(f"JSON解析エラー: {str(e)}")
        sys.exit(1)


def main():
    """メイン処理"""
    if len(sys.argv) < 4:
        print("使用方法:")
        print("  python fill_template.py <template_path> <data_json> <output_path>")
        print()
        print("例:")
        print('  python fill_template.py "11-1.xlsx" \'{"contract_date": {"cell": "C3", "value": "2025-03-01", "type": "date"}}\' "output.xlsx"')
        sys.exit(1)

    template_path = sys.argv[1]
    json_str = sys.argv[2]
    output_path = sys.argv[3]

    # 設定を読み込む
    config = load_config_from_json(json_str)

    # テンプレートを開く
    try:
        filler = RabbynetContractFiller(template_path)
    except FileNotFoundError as e:
        print(f"エラー: {str(e)}")
        sys.exit(1)

    # 設定に基づいてセルを記入
    result = filler.fill_from_config(config)

    # 結果を表示
    print(f"記入完了: {result['success']}件成功, {result['failure']}件失敗")
    if result['errors']:
        print("\nエラー:")
        for error in result['errors']:
            print(f"  - {error}")

    # ファイルを保存
    if filler.save(output_path):
        print(f"\n✓ ファイルを保存しました: {output_path}")
    else:
        print(f"\n✗ ファイル保存に失敗しました")
        for error in filler.errors:
            print(f"  - {error}")
        sys.exit(1)


if __name__ == '__main__':
    main()
